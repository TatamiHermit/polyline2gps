import requests
import math
import openpyxl
import pandas as pd
import numpy as np
import re
from geopy.distance import geodesic
import os
import time
import datetime
import logging
from functools import wraps
import webbrowser
import json
# import modin.padas as pd

# filename = "高德地图数据采集.xlsx"
# path = os.getcwd()
# file_path = os.path.join(path, filename)





def func_time(f):
    """
    记录执行时间
    :param f:
    :return:
    """
    @wraps(f)
    def wrapper(*args, **kwargs):
        start_time: float = time.time()
        result = f(*args, **kwargs)
        end_time: float = time.time()
        logger.info(f'{f.__name__}总计用时{round((end_time - start_time),2)}')
        return result
    return wrapper


@func_time
def open_gps_html(filename):
    with open("gps.js", "r") as f:
        data_func = f.read()
    with open(filename, "w") as f:
        f.write(''.join([
            '<html>',
            '<head>',
            '<meta charset="utf-8"> ',
            '<meta http-equiv="X-UA-Compatible" content="IE=edge">',
            '<meta name="viewport" content="initial-scale=1.0, user-scalable=no, width=device-width">',
            '<title>Polyline Replay</title>',
            '<link rel="stylesheet" href="https://a.amap.com/jsapi_demos/static/demo-center/css/demo-center.css"/>  ',
            '<link href="gps.css" rel="stylesheet" type="text/css"/>  ',
            '</head>  ',
            '<body>  ',
            '<div id="container"></div>',
            '<div class="input-card">',
            '<h4>Polyline Replay</h4>',
            '<div class="input-item">',
            '<input type="button" class="btn" value="Replay" id="start" onclick="startAnimation()"/>',
            '<input type="button" class="btn" value="Pause" id="pause" onclick="pauseAnimation()"/>',
            '</div>',
            '<div class="input-item">',
            '<input type="button" class="btn" value="Continue" id="resume" onclick="resumeAnimation()"/>',
            '<input type="button" class="btn" value="Stop" id="stop" onclick="stopAnimation()"/>',
            '</div>',
            '</div>',
            '<script type="text/javascript" src="https://webapi.amap.com/maps?v=1.4.15&key=您申请的key值"></script>',
            '<script>',
            line_array,
            data_func,
            vs_js,
            '</script>',
            '</body>',
            '</html>'
        ]))
    webbrowser.open_new_tab(filename)

def write_log(name):
    global logger
    logger = logging.getLogger(name)
    logger.setLevel(logging.INFO)

    fh = logging.FileHandler(f'{name}{time.strftime("%Y-%m-%d-%H-%M-%S")}.log')
    fh.setLevel(logging.INFO)

    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)

    formatter = logging.Formatter(
        '[%(asctime)s][%(levelname)s] ## %(message)s')
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)

    logger.addHandler(fh)
    logger.addHandler(ch)


def gcj2wgs(location):
    # location格式如下：locations[1] = "113.923745,22.530824"
    lon = float(location[0:location.find(",")])
    lat = float(location[location.find(",") + 1:len(location)])
    a = 6378245.0  # 克拉索夫斯基椭球参数长半轴a
    ee = 0.00669342162296594323  # 克拉索夫斯基椭球参数第一偏心率平方
    PI = 3.14159265358979324  # 圆周率
    # 以下为转换公式
    x = lon - 105.0
    y = lat - 35.0
    # 经度
    dLon = 300.0 + x + 2.0 * y + 0.1 * x * x + \
        0.1 * x * y + 0.1 * math.sqrt(abs(x))
    dLon += (20.0 * math.sin(6.0 * x * PI) + 20.0 *
             math.sin(2.0 * x * PI)) * 2.0 / 3.0
    dLon += (20.0 * math.sin(x * PI) + 40.0 *
             math.sin(x / 3.0 * PI)) * 2.0 / 3.0
    dLon += (150.0 * math.sin(x / 12.0 * PI) + 300.0 *
             math.sin(x / 30.0 * PI)) * 2.0 / 3.0
    # 纬度
    dLat = -100.0 + 2.0 * x + 3.0 * y + 0.2 * y * \
        y + 0.1 * x * y + 0.2 * math.sqrt(abs(x))
    dLat += (20.0 * math.sin(6.0 * x * PI) + 20.0 *
             math.sin(2.0 * x * PI)) * 2.0 / 3.0
    dLat += (20.0 * math.sin(y * PI) + 40.0 *
             math.sin(y / 3.0 * PI)) * 2.0 / 3.0
    dLat += (160.0 * math.sin(y / 12.0 * PI) + 320 *
             math.sin(y * PI / 30.0)) * 2.0 / 3.0
    radLat = lat / 180.0 * PI
    magic = math.sin(radLat)
    magic = 1 - ee * magic * magic
    sqrt_magic = math.sqrt(magic)
    dLat = (dLat * 180.0) / ((a * (1 - ee)) / (magic * sqrt_magic) * PI)
    dLon = (dLon * 180.0) / (a / sqrt_magic * math.cos(radLat) * PI)
    wgsLon = lon - dLon
    wgsLat = lat - dLat
    return wgsLon, wgsLat


# 将地址转化成坐标
@func_time
def get_loc(address):
    parameters = {
        'key': '429a22f69c6320aac18b2aa9b2aef883',
        'address': address}
    base = 'https://restapi.amap.com/v3/geocode/geo'
    response = requests.get(base, parameters)
    answer = response.json()
    lon = answer['geocodes'][0]['location'].split(',')[0]
    lat = answer['geocodes'][0]['location'].split(',')[1]
    return lon + "," + lat


# 获取polyline坐标串接口：https://restapi.amap.com/v3/direction/driving?origin=116.45925,39.910031&destination=116.587922,40.081577&output=xml&key=429a22f69c6320aac18b2aa9b2aef883
@func_time
def get_polyline(src, des):
    """
    :param src:始发地
    :param des:目的地
    :return:
    """
    origin = get_loc(src)
    destination = get_loc(des)
    base_url = "https://restapi.amap.com/v3/direction/driving?"
    parameter = {
        'origin': origin,
        'destination': destination,
        'key': '429a22f69c6320aac18b2aa9b2aef883',
        'strategy': strategy}
    res = requests.get(base_url, parameter)
    json = res.json()
    # 获取steps字段
    steps = json['route']['paths'][0]['steps']
    # # 写入excel
    # book = openpyxl.Workbook()
    # ws = book.active  # ws操作sheet页
    # sheet = book.create_sheet('Polyline', 0)
    # for i, j in enumerate(steps):
    #     pol = steps[i]['polyline']
    #     sheet.cell(i + 1, 1).value = 'polyline'
    #     sheet.cell(i + 1, 2).value = pol
    # book.save(filename)
    # book.close()
    polyline_list = []
    pol_list = [steps[i]['polyline'] for i in range(len(steps))]
    for pol in pol_list:
        lst = pol.split(';')
        for i in lst:
            polyline_list.append(i)
    return polyline_list


@func_time
def gcjLocation_to_wgs(src,des):
    wgsLocation = []
    gcjLocation = get_polyline(src,des)
    wgsLocation = [str(gcj2wgs(gcjLocation[i])).replace("(", "").replace(")", "") for i in range(len(gcjLocation))]
    return wgsLocation


#将坐标拆分成经纬度
@func_time
def save_location_to_excel(src,des):
    global wgs84_df
    wgsLocation = gcjLocation_to_wgs(src,des)
    df = pd.DataFrame(wgsLocation,columns=['wgsLocation']).drop_duplicates()
    wgs84_df = df['wgsLocation'].str.split(',',expand = True).rename(columns={0:'lon_wgs84',1:'lat_wgs84'})
    wgs84_df.to_excel(file_path ,index=False)


# 转成十六进制
def to_hex(wgs):
    msarc = wgs * 3600 * 1000
    hexarc = hex(int(msarc)).upper()
    hexarc2 = hexarc[2:]
    output0 = '0' * (8 - len(hexarc2)) + hexarc2
    output = output0[:2] + ' ' + output0[2:4] + \
        ' ' + output0[4:6] + ' ' + output0[6:] + ' '
    return output

@func_time
def save_polyline_to_excel(src,des):
    """
    保存原始经纬度坐标至excel
    :param src:
    :param des:
    :return:
    """
    # global mergedf
    polyline = get_polyline(src, des)
    df = pd.DataFrame(polyline, columns=['polyline']).drop_duplicates()
    orig_df = df['polyline'].str.split(',', expand=True).rename(columns={0: 'orig_lon', 1: 'orig_lat'})
    # mergedf = pd.concat([wgs84_df, orig_df], axis=1).fillna('')
    # print(mergedf)
    orig_df.to_excel(orig_file, index=False)

# # 将获取的gcjLocation数据排成一列
# @func_time
# # def save_poly_to_excel():  # 保存坐标串数据
# #     coordinates = []
# #     wb = openpyxl.load_workbook(file_path)
#     ws = wb['Polyline']
#     second_column = ws['B']
#     for x in range(len(second_column)):
#         val = second_column[x].value
#         poi = val.split(';')
#         for i in poi:
#             coordinates.append(i)
#     ws1 = wb.create_sheet("Coordinates")
#     ws1.cell(1, 1).value = "gcjLocation"     # 或写成ws1['A1'] = "gcjLocation"
#     ws1.column_dimensions['A'].width = 30  # 设置A列列宽
#     for k, v in enumerate(coordinates):
#         ws1.cell(k + 2, 1).value = v  # 数据写入excel
#     ws1.cell(1, 2).value = "wgsLocation"  # 或写成ws1['A1'] = "gcjLocation"
#     ws1.column_dimensions['B'].width = 30
#     for x, y in enumerate(coordinates):
#         ws1.cell(x +2,2).value = str(gcj2wgs(y)).replace("(", "").replace(")", "")
#     df = pd.DataFrame(ws1.values).drop_duplicates()  # 数据去重
#     df.to_excel(file_path, sheet_name="Coordinates", index=False, header=False)
#     # wb.save(file_path)
#
#
# # 将坐标差分成经纬度
#
# @func_time
# def split_data():
#     lon_wgs84 = []
#     lat_wgs84 = []
#     global col
#     wb = openpyxl.load_workbook(file_path)
#     ws1 = wb['Coordinates']
#     ws1.cell(1, 3).value = "lon_wgs84"
#     ws1.cell(1, 4).value = "lat_wgs84"
#     col = ws1['B']
#     for x in range(1, len(col)):
#         vals = col[x].value
#         datas = vals.split(',')
#         for index, data in enumerate(datas):
#             if index == 0:
#                 lon_wgs84.append(data)
#             else:
#                 lat_wgs84.append(data)
#     for a, b in enumerate(lon_wgs84):
#         ws1.cell(a + 2, 3).value = b  # 数据写入excel
#     for c, d in enumerate(lat_wgs84):
#         ws1.cell(c + 2, 4).value = d
#     # return lon_wgs84,lat_wgs84
#     wb.save(file_path)
#     wb.close()



def interpolation(loc1, loc2, a, b, vs_h):
    """
    对两个坐标点之间线性插值
    :param loc1: 坐标1
    :param loc2: 坐标2
    :param a: 坐标1的经/纬度
    :param b: 坐标2的经/纬度
    :param vs_h: 车速(km/h)
    :return: 返回插值后的坐标
    """
    n = (geodesic(loc1, loc2).m) / (vs_h * 1000 * bus_interval_time / 3600)  # m/100ms
    delta = abs((a - b) / n)  # 步长
    if a > b:
        interValue = np.arange(a, b, -delta)
    elif a<b:
        interValue = np.arange(a, b, delta)
    elif a==b:
        interValue = np.linspace(a,b,n+1)
    return interValue


def interval_after_orig_lonlat():  # 将插值后的原始经纬度值写入表中
    interOrigLonList = []
    interOrigLatList = []
    # df = pd.read_excel(file_path).drop_duplicates().head(10)
    df = pd.read_excel(orig_file).drop_duplicates()
    orig_lat = pd.Series(df['orig_lat']).tolist()
    orig_lon = pd.Series(df['orig_lon']).tolist()
    orig_loc = list(zip(orig_lat, orig_lon))
    for i in range(len(orig_loc)):
        if i < len(orig_loc) - 1:
            interOrigLon = interpolation(
                orig_loc[i], orig_loc[i + 1], orig_lon[i], orig_lon[i + 1], vs_h)
            interOrigLonList.extend(pd.Series(interOrigLon))
            interOrigLat = interpolation(
                orig_loc[i], orig_loc[i + 1], orig_lat[i], orig_lat[i + 1], vs_h)
            interOrigLatList.extend(pd.Series(interOrigLat))
        else:
            break
    df.drop(df.columns[:], axis=1)
    df = pd.concat([pd.DataFrame({'origLonValues': interOrigLonList}), pd.DataFrame(
        {'origLatValues': interOrigLatList})], axis=1)
    df.to_excel(orig_file, sheet_name="Coordinates", index=False)


@func_time
def get_linearr(): #获取lineArr数组
    global line_array, vs_js
    df0 = pd.read_excel(orig_file)
    lonlist = df0['origLonValues'].tolist()
    latlist = df0['origLatValues'].tolist()
    lineArr = [[lonlist[i], latlist[i]] for i in range(len(lonlist))]
    linestr = [(str(i)+',') for i in lineArr]
    a = ''
    b = a.join(linestr)
    # print(b)
    line_array = ('var lineArr = '+ '['+ b[:-1]+ '];')
    # print(line_array)
    # "var map = new AMap.Map("container", {resizeEnable: true,center: ["+ linestr[0] + "],zoom: 17});"
    vs_js = "function startAnimation () { marker.moveAlong(lineArr, "+ str(vs_h) +");}"
    # marker = new AMap.Marker({map: map,position: [116.478935, 39.997761],icon: "https://webapi.amap.com/images/car.png",offset: new AMap.Pixel(-26, -13),autoRotation: true,angle: -90,});


@func_time
def interval_after_lonlat():  # 将插值后的wgs84格式经纬度值写入表中
    interLonList = []
    interLatList = []
    global df
    # df = pd.read_excel(file_path).drop_duplicates().head(10)
    df = pd.read_excel(file_path).drop_duplicates()
    lat = pd.Series(df['lat_wgs84']).tolist()
    lon = pd.Series(df['lon_wgs84']).tolist()
    loc = list(zip(lat, lon))
    for i in range(len(loc)):
        if i < len(loc) - 1:
            interLon = interpolation(
                loc[i], loc[i + 1], lon[i], lon[i + 1], vs_h)
            interLonList.extend(pd.Series(interLon))
            interLat = interpolation(
                loc[i], loc[i + 1], lat[i], lat[i + 1], vs_h)
            interLatList.extend(pd.Series(interLat))
        else:
            break
    df.drop(df.columns[:], axis=1)
    df = pd.concat([pd.DataFrame({'lonValues': interLonList}), pd.DataFrame(
        {'latValues': interLatList})], axis=1)
    df.to_excel(file_path, sheet_name="Coordinates", index=False)


def series(lst, length):
    """
    生成burstID序列，序列长度等于各列长
    :param lst: 传[0, 8, 16, 24] 或 [0,1,2,3]
    :param length: 生成序列的长度
    :return: 返回生成的序列
    """
    # 生成dec_byte6序列，列表形式 birst_id
    # a = [0, 8, 16, 24]
    a = []
    for j in range(length):
        if j <= int(length / 4):
            for i in lst:
                a.append(i)
        else:
            break
    return a[0:length]


def heading_angle(lon_a, lat_a, lon_b, lat_b):  # 计算车辆的航向角
    y = math.sin(lon_b - lon_a) * math.cos(lat_b)
    x = math.cos(lat_a) * math.sin(lat_b) - math.sin(lat_a) * \
        math.cos(lat_b) * math.cos(lon_b - lon_a)
    # angle = math.atan2(y,x)
    angle = round(math.degrees(math.atan2(y, x)))
    if angle < 0:
        angle = angle + 360
    return angle


@func_time
def angle_list():  # 航向角列表
    angle = []
    latList = df['lonValues'].values.tolist()
    lonList = df['latValues'].values.tolist()
    for i in range(len(latList)):
        if i < len(latList) - 1:
            angle.append(heading_angle(lonList[i], latList[i], lonList[i + 1], latList[i + 1]))
    angle.append(angle[i - 1])
    # print(len(angle))
    return angle


@func_time
def datetime_to_msg(can_id):
    global date
    brst_ids = [0, 1, 2, 3]
    # brst_ids = [3, 0, 1, 2]
    # seq = 1000000
    id = int(int(can_id, 16) / 8)
    time_list = pd.date_range(
        basetime,
        freq="100ms",
        periods=len(df)).strftime('%Y-%m-%d-%H-%M-%S-%f')
    a = time_list.str.split('-', expand=True)
    b = pd.Series(5., a)
    date = b.reset_index()
    date.columns = ['YY', 'MM', 'DD', 'H', 'M', 'S', 'MS', '0']
    date['YY'] = date['YY'].str[2:]
    date['MS'] = date['MS'].str[:3]
    date['ID'] = series(brst_ids, len(df))
    date['BIN-YY'] = date['YY'].apply(lambda x: '{:09b}'.format(int(x)))
    date['BIN-MM'] = date['MM'].apply(lambda x: '{:05b}'.format(int(x)))
    date['BIN-DD'] = date['DD'].apply(lambda x: '{:06b}'.format(int(x)))
    date['BIN-H'] = date['H'].apply(lambda x: '{:06b}'.format(int(x)))
    date['BIN-M'] = date['M'].apply(lambda x: '{:07b}'.format(int(x)))
    date['BIN-S'] = date['S'].apply(lambda x: '{:07b}'.format(int(x)))
    date['BIN-MS'] = date['MS'].apply(lambda x: '{:011b}'.format(int(x)))
    date['BIN-ID'] = date['ID'].apply(lambda x: '{:02b}'.format(int(x)))
    date['BIN-53'] = date['BIN-YY']+date['BIN-MM']+date['BIN-DD']+date['BIN-H']+date['BIN-M']+date['BIN-S']+date['BIN-MS']+date['BIN-ID']
    date['B0'] = date['BIN-53'].str[0:8]
    date['B1'] = date['BIN-53'].str[8:16]
    date['B2'] = date['BIN-53'].str[16:24]
    date['B3'] = date['BIN-53'].str[24:32]
    date['B4'] = date['BIN-53'].str[32:40]
    date['B5'] = date['BIN-53'].str[40:48]
    date['B6-temp'] = date['BIN-53'].str[48:53] + '000'
    # checksum = int(byte0, 2) + int(byte1, 2) + int(byte2, 2) + int(byte3, 2) + int(byte4, 2) + int(byte5, 2) + id + byte6                                                                                        2) + id + byte6
    date['B0-int'] = date['B0'].apply(lambda x: int(x, 2))
    date['B1-int'] = date['B1'].apply(lambda x: int(x, 2))
    date['B2-int'] = date['B2'].apply(lambda x: int(x, 2))
    date['B3-int'] = date['B3'].apply(lambda x: int(x, 2))
    date['B4-int'] = date['B4'].apply(lambda x: int(x, 2))
    date['B5-int'] = date['B5'].apply(lambda x: int(x, 2))
    date['B6-temp-int'] = date['B6-temp'].apply(lambda x: int(x, 2))
    date['B6&F8'] = np.bitwise_and(date['B6-temp-int'], 248)
    date['B6&F8 >>3'] = np.right_shift(date['B6&F8'], 3)
    date['CHKSUM'] = date['B0-int']+date['B1-int']+date['B2-int']+date['B3-int']+date['B4-int']+date['B5-int']+date['B6&F8 >>3']+id
    date['BIN-CHKSUM'] = date['CHKSUM'].apply(lambda x: '{:011b}'.format(int(x)))
    date['BIN-64'] = date['BIN-53']+date['BIN-CHKSUM']
    date['B6'] = date['BIN-64'].str[48:56]
    date['B7'] = date['BIN-64'].str[56:64]
    date['B6-int'] = date['B6'].apply(lambda x: int(x, 2))
    date['B7-int'] = date['B7'].apply(lambda x: int(x, 2))
    date['B0-hex'] = date['B0-int'].apply(lambda x: '{0:02X}'.format(x))
    date['B1-hex'] = date['B1-int'].apply(lambda x: '{0:02X}'.format(x))
    date['B2-hex'] = date['B2-int'].apply(lambda x: '{0:02X}'.format(x))
    date['B3-hex'] = date['B3-int'].apply(lambda x: '{0:02X}'.format(x))
    date['B4-hex'] = date['B4-int'].apply(lambda x: '{0:02X}'.format(x))
    date['B5-hex'] = date['B5-int'].apply(lambda x: '{0:02X}'.format(x))
    date['B6-hex'] = date['B6-int'].apply(lambda x: '{0:02X}'.format(x))
    date['B7-hex'] = date['B7-int'].apply(lambda x: '{0:02X}'.format(x))
    date['time'] = [format((0.001 + i / 10), '0.6f') for i in range(len(df))]
    date['channel'] = 0
    date['asc-msg'] = date['time'] + " 0 " + can_id + '             Rx  d 8 ' + \
                     date['B0-hex'] + ' ' + date['B1-hex'] + ' ' + date['B2-hex'] + ' ' + date['B3-hex'] + ' ' \
                     + date['B4-hex'] + ' ' + date['B5-hex'] + ' ' + date['B6-hex'] + ' ' + date['B7-hex'] + ' '
    # print(date['asc-msg'])
    # ang.to_csv("date.csv", index = False)  # 按指定列名顺序输出df
    return date[['time', 'asc-msg']]


@func_time
def angle_to_msg(can_id):
    """
    :param can_id: 351
    :return: msg: 十六进制的报文
    """
    global ang
    msg = []
    brst_ids = [0, 1, 2, 3]
    id = int(int(can_id, 16) / 8)
    ang = pd.DataFrame()
    ang['angle'] = angle_list()
    ang['speed'] = vs_h
    ang['elevation'] = elevation+1000000
    ang['ID'] = series(brst_ids, len(df))
    ang['blank'] = 0
    ang['BIN-angle'] = ang['angle'].apply(lambda x: '{:013b}'.format((int(x) + 90) * 10))
    ang['BIN-speed'] = ang['speed'].apply(lambda x: '{:08b}'.format(int(x))) + '0'
    ang['BIN-elevation'] = ang['elevation'].apply(lambda x: '{:023b}'.format(int(x)))
    ang['BIN-ID'] = ang['ID'].apply(lambda x: '{:06b}'.format(int(x)))
    ang['BIN-blank'] = ang['blank'].apply(lambda x: '{:03b}'.format(int(x)))
    ang['BIN-53'] = ang['BIN-angle'] + ang['BIN-speed'] + ang['BIN-elevation'] + ang['BIN-ID'] + ang['BIN-blank']
    ang['B0'] = ang['BIN-53'].str[0:8]
    ang['B1'] = ang['BIN-53'].str[8:16]
    ang['B2'] = ang['BIN-53'].str[16:24]
    ang['B3'] = ang['BIN-53'].str[24:32]
    ang['B4'] = ang['BIN-53'].str[32:40]
    ang['B5'] = ang['BIN-53'].str[40:48]
    ang['B6-temp'] = ang['BIN-53'].str[48:53] + '000'
    ang['B0-int'] = ang['B0'].apply(lambda x: int(x, 2))
    ang['B1-int'] = ang['B1'].apply(lambda x: int(x, 2))
    ang['B2-int'] = ang['B2'].apply(lambda x: int(x, 2))
    ang['B3-int'] = ang['B3'].apply(lambda x: int(x, 2))
    ang['B4-int'] = ang['B4'].apply(lambda x: int(x, 2))
    ang['B5-int'] = ang['B5'].apply(lambda x: int(x, 2))
    ang['B6-temp-int'] = ang['B6-temp'].apply(lambda x: int(x, 2))
    ang['B6&F8'] = np.bitwise_and(ang['B6-temp-int'], 248)  # 248=0xF8
    ang['B6&F8 >>3'] = np.right_shift(ang['B6&F8'], 3)
    ang['CHKSUM'] = ang['B0-int']+ang['B1-int']+ang['B2-int']+ang['B3-int']+ang['B4-int']+ang['B5-int']+ang['B6&F8 >>3']+id
    ang['BIN-CHKSUM'] = ang['CHKSUM'].apply(lambda x: '{:011b}'.format(int(x)))
    ang['BIN-64'] = ang['BIN-53']+ang['BIN-CHKSUM']
    ang['B6'] = ang['BIN-64'].str[48:56]
    ang['B7'] = ang['BIN-64'].str[56:64]
    ang['B6-int'] = ang['B6'].apply(lambda x: int(x, 2))
    ang['B7-int'] = ang['B7'].apply(lambda x: int(x, 2))
    ang['B0-hex'] = ang['B0-int'].apply(lambda x: '{0:02X}'.format(x))
    ang['B1-hex'] = ang['B1-int'].apply(lambda x: '{0:02X}'.format(x))
    ang['B2-hex'] = ang['B2-int'].apply(lambda x: '{0:02X}'.format(x))
    ang['B3-hex'] = ang['B3-int'].apply(lambda x: '{0:02X}'.format(x))
    ang['B4-hex'] = ang['B4-int'].apply(lambda x: '{0:02X}'.format(x))
    ang['B5-hex'] = ang['B5-int'].apply(lambda x: '{0:02X}'.format(x))
    ang['B6-hex'] = ang['B6-int'].apply(lambda x: '{0:02X}'.format(x))
    ang['B7-hex'] = ang['B7-int'].apply(lambda x: '{0:02X}'.format(x))
    ang['time'] = [format((0.000 + i / 10), '0.6f') for i in range(len(df))]
    ang['channel'] = 0
    ang['asc-msg'] = ang['time'] + " 0 " + can_id + '             Rx  d 8 ' + \
                     ang['B0-hex'] + ' ' + ang['B1-hex'] + ' ' + ang['B2-hex'] + ' ' + ang['B3-hex'] + ' ' \
                     + ang['B4-hex'] + ' ' + ang['B5-hex'] + ' ' + ang['B6-hex'] + ' ' + ang['B7-hex'] + ' '
    # print(ang['asc-msg'])
    # ang.to_csv("ang.csv", index = False)  # 按指定列名顺序输出df
    return ang[['time', 'asc-msg']]

@func_time
def lat_to_msg(can_id):
    """
    :param data: 传lat_hex_lon列的hex值
    :param can_id: 纬度为'35C'，经度为'35D'
    :return: msg: 十六进制的报文
    """
    global lat
    msg = []
    brst_ids = [0, 1, 2, 3]
    id = int(int(can_id, 16) / 8)
    # print('数据总长度为', len(df))
    lat = pd.DataFrame()
    lat['BIN-lon'] = df['latValues'].apply(lambda x: '{:032b}'.format(int(x*3600*1000)))
    lat['BIN-spare'] = '0000000000000000' #16位
    lat['ID'] = series(brst_ids, len(df))
    lat['BIN-ID'] = lat['ID'].apply(lambda x: '{:05b}'.format(int(x)))
    lat['BIN-53'] = lat['BIN-lon'] + lat['BIN-spare'] + lat['BIN-ID']
    lat['B0'] = lat['BIN-53'].str[0:8]
    lat['B1'] = lat['BIN-53'].str[8:16]
    lat['B2'] = lat['BIN-53'].str[16:24]
    lat['B3'] = lat['BIN-53'].str[24:32]
    lat['B4'] = lat['BIN-53'].str[32:40]
    lat['B5'] = lat['BIN-53'].str[40:48]
    lat['B6-temp'] = lat['BIN-53'].str[48:53] + '000'
    lat['B0-int'] = lat['B0'].apply(lambda x: int(x, 2))
    lat['B1-int'] = lat['B1'].apply(lambda x: int(x, 2))
    lat['B2-int'] = lat['B2'].apply(lambda x: int(x, 2))
    lat['B3-int'] = lat['B3'].apply(lambda x: int(x, 2))
    lat['B4-int'] = lat['B4'].apply(lambda x: int(x, 2))
    lat['B5-int'] = lat['B5'].apply(lambda x: int(x, 2))
    lat['B6-temp-int'] = lat['B6-temp'].apply(lambda x: int(x, 2))
    lat['B6&F8'] = np.bitwise_and(lat['B6-temp-int'], 248)
    lat['B6&F8 >>3'] = np.right_shift(lat['B6&F8'], 3)
    lat['CHKSUM'] = lat['B0-int']+lat['B1-int']+lat['B2-int']+lat['B3-int']+lat['B4-int']+lat['B5-int']+lat['B6&F8 >>3']+id
    lat['BIN-CHKSUM'] = lat['CHKSUM'].apply(lambda x: '{:011b}'.format(int(x)))
    lat['BIN-64'] = lat['BIN-53']+lat['BIN-CHKSUM']
    lat['B6'] = lat['BIN-64'].str[48:56]
    lat['B7'] = lat['BIN-64'].str[56:64]
    lat['B6-int'] = lat['B6'].apply(lambda x: int(x, 2))
    lat['B7-int'] = lat['B7'].apply(lambda x: int(x, 2))
    lat['B0-hex'] = lat['B0-int'].apply(lambda x: '{0:02X}'.format(x))
    lat['B1-hex'] = lat['B1-int'].apply(lambda x: '{0:02X}'.format(x))
    lat['B2-hex'] = lat['B2-int'].apply(lambda x: '{0:02X}'.format(x))
    lat['B3-hex'] = lat['B3-int'].apply(lambda x: '{0:02X}'.format(x))
    lat['B4-hex'] = lat['B4-int'].apply(lambda x: '{0:02X}'.format(x))
    lat['B5-hex'] = lat['B5-int'].apply(lambda x: '{0:02X}'.format(x))
    lat['B6-hex'] = lat['B6-int'].apply(lambda x: '{0:02X}'.format(x))
    lat['B7-hex'] = lat['B7-int'].apply(lambda x: '{0:02X}'.format(x))
    lat['time'] = [format((0.002 + i / 10),'0.6f') for i in range(len(df))]
    lat['channel'] = 0
    lat['asc-msg'] = lat['time'] + " 0 " + can_id + '             Rx  d 8 ' + \
                     lat['B0-hex'] + ' ' + lat['B1-hex'] + ' ' + lat['B2-hex'] + ' ' + lat['B3-hex'] + ' ' \
                     + lat['B4-hex'] + ' ' + lat['B5-hex'] + ' ' + lat['B6-hex'] + ' ' + lat['B7-hex'] + ' '
    # print(lat['asc-msg'])
    # lat.to_csv("lat.csv", index=False)  # 按指定列名顺序输出df
    return lat[['time', 'asc-msg']]

@func_time
def lon_to_msg(can_id):
    """
    :param data: 传lat_hex_lon列的hex值
    :param can_id: lat纬度为'35C'，lon经度为'35D'
    :return: msg: 十六进制的报文
    """
    global lon
    msg = []
    brst_ids = [0, 8, 16, 24]
    id = int(int(can_id, 16) / 8)
    # print('数据总长度为', len(df))
    lon = pd.DataFrame()
    lon['BIN-lon'] = df['lonValues'].apply(lambda x: '{:032b}'.format(int(x*3600*1000)))
    lon['BIN-spare'] = '0000000000000000' #16位
    lon['ID'] = series(brst_ids, len(df))
    lon['BIN-ID'] = lon['ID'].apply(lambda x: '{:05b}'.format(int(x)))
    lon['BIN-53'] = lon['BIN-lon'] + lon['BIN-spare'] + lon['BIN-ID']
    lon['B0'] = lon['BIN-53'].str[0:8]
    lon['B1'] = lon['BIN-53'].str[8:16]
    lon['B2'] = lon['BIN-53'].str[16:24]
    lon['B3'] = lon['BIN-53'].str[24:32]
    lon['B4'] = lon['BIN-53'].str[32:40]
    lon['B5'] = lon['BIN-53'].str[40:48]
    lon['B6-temp'] = lon['BIN-53'].str[48:53] + '000'
    lon['B0-int'] = lon['B0'].apply(lambda x: int(x, 2))
    lon['B1-int'] = lon['B1'].apply(lambda x: int(x, 2))
    lon['B2-int'] = lon['B2'].apply(lambda x: int(x, 2))
    lon['B3-int'] = lon['B3'].apply(lambda x: int(x, 2))
    lon['B4-int'] = lon['B4'].apply(lambda x: int(x, 2))
    lon['B5-int'] = lon['B5'].apply(lambda x: int(x, 2))
    lon['B6-temp-int'] = lon['B6-temp'].apply(lambda x: int(x, 2))
    lon['B6&F8'] = np.bitwise_and(lon['B6-temp-int'], 248)
    lon['B6&F8 >>3'] = np.right_shift(lon['B6&F8'], 3)
    lon['CHKSUM'] = lon['B0-int']+lon['B1-int']+lon['B2-int']+lon['B3-int']+lon['B4-int']+lon['B5-int']+lon['B6&F8 >>3']+id
    lon['BIN-CHKSUM'] = lon['CHKSUM'].apply(lambda x: '{:011b}'.format(int(x)))
    lon['BIN-64'] = lon['BIN-53']+lon['BIN-CHKSUM']
    lon['B6'] = lon['BIN-64'].str[48:56]
    lon['B7'] = lon['BIN-64'].str[56:64]
    lon['B6-int'] = lon['B6'].apply(lambda x: int(x, 2))
    lon['B7-int'] = lon['B7'].apply(lambda x: int(x, 2))
    lon['B0-hex'] = lon['B0-int'].apply(lambda x: '{0:02X}'.format(x))
    lon['B1-hex'] = lon['B1-int'].apply(lambda x: '{0:02X}'.format(x))
    lon['B2-hex'] = lon['B2-int'].apply(lambda x: '{0:02X}'.format(x))
    lon['B3-hex'] = lon['B3-int'].apply(lambda x: '{0:02X}'.format(x))
    lon['B4-hex'] = lon['B4-int'].apply(lambda x: '{0:02X}'.format(x))
    lon['B5-hex'] = lon['B5-int'].apply(lambda x: '{0:02X}'.format(x))
    lon['B6-hex'] = lon['B6-int'].apply(lambda x: '{0:02X}'.format(x))
    lon['B7-hex'] = lon['B7-int'].apply(lambda x: '{0:02X}'.format(x))
    lon['time'] = [format((0.003 + i / 10), '0.6f') for i in range(len(df))]
    lon['channel'] = 0
    lon['asc-msg'] = lon['time'] + " 0 " + can_id + '             Rx  d 8 ' + \
                     lon['B0-hex'] + ' ' + lon['B1-hex'] + ' ' + lon['B2-hex'] + ' ' + lon['B3-hex'] + ' ' \
                     + lon['B4-hex'] + ' ' + lon['B5-hex'] + ' ' + lon['B6-hex'] + ' ' + lon['B7-hex'] + ' '
    # print(lon['asc-msg'])
    # ang.to_csv("lon.csv", index = False)  # 按指定列名顺序输出df
    return lon[['time', 'asc-msg']]


@func_time
def merge_message():
    # 删除其他列，合并Message_lat 和Message_lon
    global msg_all
    msg_all = pd.DataFrame()
    msg_all = pd.concat([date[['time', 'asc-msg']], ang[['time', 'asc-msg']], lat[['time', 'asc-msg']],
                         lon[['time', 'asc-msg']]])
    msg_all['time'] = msg_all['time'].astype('float')
    msg_all.sort_values(by=['time'], ascending=True, inplace=True)
    # print(msg_all)

@func_time
def msg_to_asc():
    msg_file = f'{time.strftime("%Y-%m-%d-%H-%M")}-{src}-{des}-{vs_h}.asc'
    with open(msg_file, 'w') as fw:
        string = "date {0} \nbase hex  timestamps absolute \nno internal events logged".format(
            datetime.datetime.now().ctime())
        fw.write(string)
        fw.write('\n')
        lst = msg_all['asc-msg'].values.tolist()
        for i, signal in enumerate(lst):
            fw.write(lst[i])
            fw.write('\n')


if __name__ == "__main__":
    orig_file = "原始坐标.xlsx"
    filename = "高德地图数据采集.xlsx"
    path = os.getcwd()
    file_path = os.path.join(path, filename)
    start = time.time()
    bus_interval_time = 0.1  # GPS信号帧间隔,单位秒
    gps_date_can_id = '351'  # GPS日期CAN ID
    gps_angle_can_id = '353'  # GPS航向角CAN ID
    gps_lat_can_id = '35C'  # GPS纬度CAN ID
    gps_lon_can_id = '35D'  # GPS经度CAN ID
    vs_h = 80000  # 车速, 单位千米/小时
    src = "上海市" + "巨峰路地铁站"  # 起点，可以是坐标点，如果是地址需要加上城市
    des = "上海市" + "虹桥国际机场"  # 终点，可以是坐标点，如果是地址需要加上城市
    strategy = '不走高速' #高速优先  躲避拥堵  不走高速  避免收费
    elevation = 400  # interval_after_orig_lonlat()
    basetime = "2020-01-17 8:05:50"  # 信号开始时间
    write_log(f'{time.strftime("%Y-%m-%d-%H-%M")}-{src}-{des}-{vs_h}')
    get_polyline(src, des)
    save_location_to_excel(src, des)
    save_polyline_to_excel(src, des)
    interval_after_orig_lonlat()
    get_linearr()
    open_gps_html("gps.html")
    interval_after_lonlat()
    # datetime_to_msg(gps_date_can_id)
    # angle_to_msg(gps_angle_can_id)
    # lon_to_msg(gps_lon_can_id)
    # lat_to_msg(gps_lat_can_id)
    # merge_message()
    # msg_to_asc()
    logger.info(f'总计用时{round((time.time() - start),4)}')
